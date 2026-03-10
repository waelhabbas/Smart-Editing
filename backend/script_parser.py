"""
Parse script files (Word .docx, Excel .xlsx, CSV, TXT) or Google Docs/Sheets URLs
into numbered shots. Supports Arabic and English shot labels.
"""

import re
import csv
import io
import tempfile
from pathlib import Path

import httpx
from docx import Document
from openpyxl import load_workbook


# Patterns to detect shot headers
SHOT_PATTERNS = [
    r"(?:shot|شوت|شات)\s*(\d+)\s*[:\.\-\s]",   # "Shot 1:", "شوت 1:", etc.
    r"^(\d+)\s*[:\.)\-]\s*",                       # "1:", "1.", "1)", "1-"
    r"#\s*(\d+)\s*[:\.)\-\s]",                      # "#1:"
]

# Google URL patterns
GOOGLE_DOC_PATTERN = re.compile(
    r"https?://docs\.google\.com/document/d/([a-zA-Z0-9_-]+)"
)
GOOGLE_SHEET_PATTERN = re.compile(
    r"https?://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)"
)


def parse_docx(file_path: str) -> str:
    """Read a Word .docx file and return its text content."""
    doc = Document(file_path)
    lines = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            lines.append(text)
    return "\n".join(lines)


def parse_txt(file_path: str) -> str:
    """Read a plain text file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def parse_xlsx(file_path: str) -> list[dict]:
    """
    Read an Excel .xlsx file. Each row = one shot.
    Single column: row number = shot number, cell = text.
    """
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    shots = []
    shot_num = 1

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Get first non-empty cell in the row
        text = None
        for cell in row:
            if cell is not None:
                text = str(cell).strip()
                break
        if text:
            shots.append({"shot_number": shot_num, "text": text})
            shot_num += 1

    wb.close()
    return shots


def parse_csv_file(file_path: str) -> list[dict]:
    """
    Read a CSV file. Each row = one shot.
    """
    shots = []
    shot_num = 1

    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            # Get first non-empty cell
            text = None
            for cell in row:
                cell = cell.strip()
                if cell:
                    text = cell
                    break
            if text:
                shots.append({"shot_number": shot_num, "text": text})
                shot_num += 1

    return shots


def parse_csv_text(text: str) -> list[dict]:
    """Parse CSV content from a string (for Google Sheets export)."""
    shots = []
    shot_num = 1

    reader = csv.reader(io.StringIO(text))
    for row in reader:
        cell_text = None
        for cell in row:
            cell = cell.strip()
            if cell:
                cell_text = cell
                break
        if cell_text:
            shots.append({"shot_number": shot_num, "text": cell_text})
            shot_num += 1

    return shots


def fetch_google_doc(url: str) -> list[dict]:
    """
    Fetch a Google Doc by URL and parse it into shots.
    The doc must be shared with "Anyone with the link".
    """
    match = GOOGLE_DOC_PATTERN.search(url)
    if not match:
        raise ValueError("رابط Google Doc غير صالح")

    doc_id = match.group(1)
    export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"

    response = httpx.get(export_url, follow_redirects=True, timeout=30)
    if response.status_code != 200:
        raise ValueError("تعذر الوصول للمستند. تأكد أنه مشترك بـ 'Anyone with the link'")

    text = response.text
    return split_into_shots(text)


def fetch_google_sheet(url: str) -> list[dict]:
    """
    Fetch a Google Sheet by URL and parse it into shots.
    Each row = one shot. The sheet must be shared with "Anyone with the link".
    """
    match = GOOGLE_SHEET_PATTERN.search(url)
    if not match:
        raise ValueError("رابط Google Sheet غير صالح")

    sheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"

    response = httpx.get(export_url, follow_redirects=True, timeout=30)
    if response.status_code != 200:
        raise ValueError("تعذر الوصول للشيت. تأكد أنه مشترك بـ 'Anyone with the link'")

    return parse_csv_text(response.text)


def is_google_url(url: str) -> bool:
    """Check if a string is a Google Docs or Sheets URL."""
    return bool(GOOGLE_DOC_PATTERN.search(url) or GOOGLE_SHEET_PATTERN.search(url))


def parse_google_url(url: str) -> list[dict]:
    """Parse a Google Docs or Sheets URL into shots."""
    if GOOGLE_DOC_PATTERN.search(url):
        return fetch_google_doc(url)
    elif GOOGLE_SHEET_PATTERN.search(url):
        return fetch_google_sheet(url)
    else:
        raise ValueError("الرابط غير معروف. استخدم رابط Google Docs أو Google Sheets")


def split_into_shots(text: str) -> list[dict]:
    """
    Split script text into numbered shots.

    Returns:
        List of dicts: [{"shot_number": int, "text": str}, ...]
    """
    lines = text.split("\n")
    shots = []
    current_shot = None
    current_text_lines = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Try to match a shot header
        shot_number = None
        remaining_text = line

        for pattern in SHOT_PATTERNS:
            match = re.match(pattern, line, re.IGNORECASE)
            if match:
                shot_number = int(match.group(1))
                remaining_text = line[match.end():].strip()
                break

        if shot_number is not None:
            # Save previous shot
            if current_shot is not None:
                shots.append({
                    "shot_number": current_shot,
                    "text": " ".join(current_text_lines).strip(),
                })

            current_shot = shot_number
            current_text_lines = [remaining_text] if remaining_text else []
        else:
            # Continue current shot or start shot 1 implicitly
            if current_shot is None:
                current_shot = 1
            current_text_lines.append(line)

    # Save last shot
    if current_shot is not None and current_text_lines:
        shots.append({
            "shot_number": current_shot,
            "text": " ".join(current_text_lines).strip(),
        })

    return shots


def parse_script(file_path: str) -> list[dict]:
    """
    Parse a script file into shots.

    Args:
        file_path: Path to .docx, .xlsx, .csv, or .txt file.

    Returns:
        List of dicts: [{"shot_number": int, "text": str}, ...]
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".docx":
        text = parse_docx(file_path)
        return split_into_shots(text)
    elif ext in (".txt", ".text"):
        text = parse_txt(file_path)
        return split_into_shots(text)
    elif ext == ".xlsx":
        return parse_xlsx(file_path)
    elif ext == ".csv":
        return parse_csv_file(file_path)
    else:
        raise ValueError(f"صيغة غير مدعومة: {ext}. استخدم .docx أو .xlsx أو .csv أو .txt")
